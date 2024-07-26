import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep

# Função para realizar a verificação dos pagamentos
def verificar_pagamentos():
    # 1 - Entrar na planilha e extrair o cpf do cliente
    try:
        planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
        pagina_clientes = planilha_clientes['Sheet1']
    except Exception as e:
        print(f"Erro ao abrir a planilha de clientes: {e}")
        return

    driver = webdriver.Chrome()
    driver.get('https://consultcpf-devaprender.netlify.app/')

    for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
        nome, valor, cpf, vencimento = linha

        # 2 - Entrar no site e usar o cpf da planilha para pesquisar o status do pagamento daquele cliente
        sleep(5)
        campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
        sleep(1)
        campo_pesquisa.clear()
        campo_pesquisa.send_keys(cpf)
        sleep(1)

        # 3 Verificar se está em dia ou atrasado
        botao_pesquisar = driver.find_element(By.XPATH, "//button[contains(@class, 'btn-custom') and contains(@class, 'btn-lg') and contains(@class, 'btn-block') and contains(@class, 'mt-3')]")
        sleep(1)
        botao_pesquisar.click()

        # Esperar até que o elemento de status esteja presente
        try:
            status = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//span[@id='statusLabel1']"))
            )

            if status.text == 'em dia':
                # 4 - Se estiver "em dia", pegar a data do pagamento e o método de pagamento 
                data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
                metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")
                data_pagamento_limpo = data_pagamento.text.split()[3]
                metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

                try:
                    # Abra a planilha para escrita
                    planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
                    pagina_fechamento = planilha_fechamento['Sheet1']
                    # Adicione a nova linha de dados
                    pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])
                    # Salve as alterações
                    planilha_fechamento.save('planilha fechamento.xlsx')
                    print(f"Adicionado {nome}, {valor}, {cpf}, {vencimento}, 'em dia', {data_pagamento_limpo}, {metodo_pagamento_limpo}")
                except Exception as e:
                    print(f"Erro ao salvar os dados na planilha de fechamento: {e}")
            else:
                # 5 - Caso contrário (se estiver atrasado), colocar o status como pendente
                try:
                    planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
                    pagina_fechamento = planilha_fechamento['Sheet1']
                    pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
                    planilha_fechamento.save('planilha fechamento.xlsx')
                    print(f"Adicionado {nome}, {valor}, {cpf}, {vencimento}, 'pendente'")
                except Exception as e:
                    print(f"Erro ao salvar os dados na planilha de fechamento: {e}")

        except Exception as e:
            print(f"Erro ao localizar o status: {e}")

    driver.quit()

# Loop infinito para manter o script em execução
while True:
    verificar_pagamentos()
    # Pequena pausa para evitar sobrecarga, ajuste conforme necessário
    sleep(1)